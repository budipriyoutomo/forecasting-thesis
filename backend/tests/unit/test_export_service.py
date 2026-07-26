"""
Fase 8 — export builder (xlsx/pdf) + ExportService. Repo/storage di-mock.
"""
import io
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

from app.services.export_service import (
    ExportService,
    build_forecast_xlsx,
    build_reorder_pdf,
    build_reorder_xlsx,
)
from app.utils.exceptions import ForbiddenRoleError, ForecastRunNotFoundError

USER = "u1"
OTHER = "u2"


def _fresult(pid="p1"):
    return SimpleNamespace(
        product_id=pid, status="COMPLETED", method_used="moving_average", selection_mode="auto",
        mape=Decimal("5.0"), mase=Decimal("0.5"),
        explanation="Moving Average dipilih.", forecast_data=[], metrics=None,
    )


def _rec(mid="m1", status="urgent"):
    return SimpleNamespace(
        material_id=mid, safety_stock=Decimal("6.6"), reorder_point=Decimal("46.6"),
        recommended_order_qty=Decimal("87"), status=status,
    )


def test_build_forecast_xlsx_bisa_dibuka_kembali():
    content = build_forecast_xlsx([_fresult("p1"), _fresult("p2")])

    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "product_id"
    assert ws.max_row == 3  # header + 2 baris


def test_build_reorder_xlsx_isi_benar():
    content = build_reorder_xlsx([_rec("m1")])

    ws = load_workbook(io.BytesIO(content)).active
    assert ws.cell(row=1, column=1).value == "material_id"
    assert ws.cell(row=2, column=5).value == "urgent"


def test_build_reorder_xlsx_kolom_eoq_tic():
    # Fase 9: kolom EOQ/biaya ditambah SETELAH status (status tetap kolom 5 — backward compat).
    rec = SimpleNamespace(
        material_id="m1", safety_stock=Decimal("6.6"), reorder_point=Decimal("46.6"),
        recommended_order_qty=Decimal("87"), status="urgent",
        buffer_stock=Decimal("10"), eoq_qty=Decimal("200"),
        ordering_cost=Decimal("100"), holding_cost=Decimal("20"), total_inventory_cost=Decimal("120"),
    )
    ws = load_workbook(io.BytesIO(build_reorder_xlsx([rec]))).active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    assert ws.cell(row=1, column=5).value == "status"  # tidak bergeser
    for col in ("buffer_stock", "eoq_qty", "ordering_cost", "holding_cost", "total_inventory_cost"):
        assert col in headers
    tic_col = headers.index("total_inventory_cost") + 1
    assert ws.cell(row=2, column=tic_col).value == 120.0


def test_build_reorder_xlsx_backward_compat_tanpa_kolom_baru():
    # rec lama (tanpa field EOQ) tetap ter-export, kolom baru kosong.
    old = SimpleNamespace(
        material_id="m1", safety_stock=Decimal("6.6"), reorder_point=Decimal("46.6"),
        recommended_order_qty=Decimal("87"), status="safe",
    )
    ws = load_workbook(io.BytesIO(build_reorder_xlsx([old]))).active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    tic_col = headers.index("total_inventory_cost") + 1
    assert ws.cell(row=2, column=tic_col).value is None


def test_build_reorder_pdf_menghasilkan_pdf():
    run = SimpleNamespace(id="r1")
    content = build_reorder_pdf(run, [_rec("m1"), _rec("m2", status="safe")])

    assert content[:4] == b"%PDF"
    assert len(content) > 100


@pytest.mark.asyncio
async def test_export_forecast_menyimpan_arsip_dan_return_bytes():
    forecast_repo = MagicMock()
    forecast_repo.get_run = _async(SimpleNamespace(id="r1", user_id=USER))
    forecast_repo.list_results = _async([_fresult("m1")])
    storage = MagicMock()
    svc = ExportService(forecast_repo, MagicMock(), storage)

    content, filename, mime = await svc.export_forecast(USER, "r1")

    assert content[:2] == b"PK"  # xlsx = zip
    assert filename == "forecast_r1.xlsx"
    storage.upload_export.assert_called_once()


@pytest.mark.asyncio
async def test_export_reorder_pdf():
    forecast_repo = MagicMock()
    forecast_repo.get_run = _async(SimpleNamespace(id="r1", user_id=USER))
    reorder_repo = MagicMock()
    reorder_repo.list_by_run = _async([_rec("m1")])
    svc = ExportService(forecast_repo, reorder_repo, storage=None)

    content, filename, mime = await svc.export_reorder(USER, "r1", "pdf")

    assert content[:4] == b"%PDF"
    assert filename == "reorder_r1.pdf"


@pytest.mark.asyncio
async def test_export_reorder_xlsx_default():
    forecast_repo = MagicMock()
    forecast_repo.get_run = _async(SimpleNamespace(id="r1", user_id=USER))
    reorder_repo = MagicMock()
    reorder_repo.list_by_run = _async([_rec("m1")])
    svc = ExportService(forecast_repo, reorder_repo, storage=None)

    content, filename, mime = await svc.export_reorder(USER, "r1", "xlsx")

    assert content[:2] == b"PK"
    assert filename == "reorder_r1.xlsx"


@pytest.mark.asyncio
async def test_export_run_tidak_ada_404():
    forecast_repo = MagicMock()
    forecast_repo.get_run = _async(None)
    svc = ExportService(forecast_repo, MagicMock(), None)

    with pytest.raises(ForecastRunNotFoundError):
        await svc.export_forecast(USER, "ghost")


@pytest.mark.asyncio
async def test_export_run_milik_user_lain_403():
    forecast_repo = MagicMock()
    forecast_repo.get_run = _async(SimpleNamespace(id="r1", user_id=OTHER))
    svc = ExportService(forecast_repo, MagicMock(), None)

    with pytest.raises(ForbiddenRoleError):
        await svc.export_forecast(USER, "r1")


@pytest.mark.asyncio
async def test_export_arsip_gagal_tidak_menggagalkan_download():
    forecast_repo = MagicMock()
    forecast_repo.get_run = _async(SimpleNamespace(id="r1", user_id=USER))
    forecast_repo.list_results = _async([_fresult("m1")])
    storage = MagicMock()
    storage.upload_export.side_effect = RuntimeError("R2 down")
    svc = ExportService(forecast_repo, MagicMock(), storage)

    content, _, _ = await svc.export_forecast(USER, "r1")

    assert content[:2] == b"PK"  # download tetap berhasil


def _async(value):
    async def _fn(*args, **kwargs):
        return value

    return _fn
