"""
ExportService — export hasil forecast & reorder ke Excel/PDF (Fase 8).

Fungsi builder MURNI (bytes, mudah dites): `build_forecast_xlsx`,
`build_reorder_xlsx`, `build_reorder_pdf`. `ExportService` mengorkestrasi:
ambil data dari repo (cek kepemilikan run), bangun file, simpan ke R2
`permanent/exports/...` (best-effort — kegagalan R2 tidak menggagalkan download),
lalu kembalikan bytes untuk diunduh.
"""
import io
from datetime import datetime, timezone

from fpdf import FPDF
from openpyxl import Workbook

from app.utils.exceptions import ForbiddenRoleError, ForecastRunNotFoundError

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"


def _num(value):
    return float(value) if value is not None else None


def build_forecast_xlsx(results) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast"
    ws.append(["product_id", "status", "method_used", "selection_mode", "mape", "mase", "explanation"])
    for r in results:
        ws.append(
            [
                str(r.product_id),
                r.status,
                r.method_used,
                r.selection_mode,
                _num(r.mape),
                _num(r.mase),
                r.explanation,
            ]
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_reorder_xlsx(recommendations) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder"
    # Kolom EOQ/biaya (Fase 9) ditambah SETELAH `status` → kolom lama tidak bergeser
    # (backward compat, larangan regresi). rec lama tanpa field baru → sel kosong.
    ws.append(
        [
            "material_id", "safety_stock", "reorder_point", "recommended_order_qty", "status",
            "buffer_stock", "eoq_qty", "ordering_cost", "holding_cost", "total_inventory_cost",
        ]
    )
    for rec in recommendations:
        ws.append(
            [
                str(rec.material_id),
                _num(rec.safety_stock),
                _num(rec.reorder_point),
                _num(rec.recommended_order_qty),
                rec.status,
                _num(getattr(rec, "buffer_stock", None)),
                _num(getattr(rec, "eoq_qty", None)),
                _num(getattr(rec, "ordering_cost", None)),
                _num(getattr(rec, "holding_cost", None)),
                _num(getattr(rec, "total_inventory_cost", None)),
            ]
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_reorder_pdf(run, recommendations) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Laporan Rekomendasi Reorder", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Run: {run.id}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0, 6, f"Dibuat: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(2)

    headers = ["Material", "Safety", "ROP", "Order qty", "Status"]
    widths = [70, 25, 25, 30, 30]
    pdf.set_font("Helvetica", "B", 9)
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, h, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 9)
    for rec in recommendations:
        row = [
            str(rec.material_id)[:32],
            f"{_num(rec.safety_stock):.2f}",
            f"{_num(rec.reorder_point):.2f}",
            f"{_num(rec.recommended_order_qty):.2f}",
            rec.status,
        ]
        for value, w in zip(row, widths):
            pdf.cell(w, 7, value, border=1)
        pdf.ln()

    out = pdf.output()
    return bytes(out)


class ExportService:
    def __init__(self, forecast_repo, reorder_repo, storage=None):
        self._forecast = forecast_repo
        self._reorder = reorder_repo
        self._storage = storage

    async def export_forecast(self, user_id: str, run_id: str):
        run = await self._require_run(user_id, run_id)
        results = await self._forecast.list_results(run_id)
        content = build_forecast_xlsx(results)
        filename = f"forecast_{run_id}.xlsx"
        self._archive(user_id, run_id, filename, content)
        return content, filename, XLSX_MIME

    async def export_reorder(self, user_id: str, run_id: str, fmt: str):
        run = await self._require_run(user_id, run_id)
        recs = await self._reorder.list_by_run(run_id)
        if fmt == "pdf":
            content = build_reorder_pdf(run, recs)
            filename, mime = f"reorder_{run_id}.pdf", PDF_MIME
        else:
            content = build_reorder_xlsx(recs)
            filename, mime = f"reorder_{run_id}.xlsx", XLSX_MIME
        self._archive(user_id, run_id, filename, content)
        return content, filename, mime

    def _archive(self, user_id: str, run_id: str, filename: str, content: bytes) -> None:
        # Simpan ke R2 permanent/exports (best-effort). Kegagalan penyimpanan
        # arsip tidak boleh menggagalkan download file oleh user.
        if self._storage is None:
            return
        try:
            self._storage.upload_export(user_id, run_id, filename, content)
        except Exception:
            pass

    async def _require_run(self, user_id: str, run_id: str):
        run = await self._forecast.get_run(run_id)
        if run is None:
            raise ForecastRunNotFoundError("Forecast run tidak ditemukan.")
        if str(run.user_id) != str(user_id):
            raise ForbiddenRoleError("Anda tidak berhak mengekspor run ini.")
        return run
